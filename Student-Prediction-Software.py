#importing the various libraries 
from tkinter import *
from tkinter import ttk, filedialog
from PIL import ImageTk, Image
from openpyxl import *
from openpyxl import *
import pathlib
import openpyxl
import pandas as pd

global Prediction
#Admin final window to read prediction result 
def admin2():

    
    #Homepage creation 
    outputw = Tk()
    outputw.title('PREDICTION OUTPUT PAGE')
    outputw['bg'] = '#29C5F6'
    outputw.resizable(False, False)  # This code helps to disable windows from resizing
    #setting window default size for home page
    window_height = 500
    window_width = 900

    screen_width = outputw.winfo_screenwidth()
    screen_height = outputw.winfo_screenheight()

    x_cordinate = int((screen_width/2) - (window_width/2))
    y_cordinate = int((screen_height/2) - (window_height/2))
    outputw.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
    
    # Frame for TreeView
    frame1 = LabelFrame(outputw, text="Excel Data", bg='#29C5F6')
    frame1.place(height=400, width=900)

    # Frame for open file dialog
    file_frame = LabelFrame(outputw, text="Open File", bg = '#29C5F6')
    file_frame.place(height=100, width=900, rely=0.8, relx=0)

    # Buttons
    button1 = Button(file_frame, text="Browse A File", command=lambda: File_dialog())
    button1.place(rely=0.65, relx=0.50)

    button2 = Button(file_frame, text="Load File", command=lambda: Load_excel_data())
    button2.place(rely=0.65, relx=0.30)

    # The file/file path text
    label_file = ttk.Label(file_frame, text="No File Selected")
    label_file.place(rely=0, relx=0)


    ## Treeview Widget
    tv1 = ttk.Treeview(frame1)
    tv1.place(relheight=1, relwidth=1) # set the height and width of the widget to 100% of its container (frame1).

    treescrolly = Scrollbar(frame1, orient="vertical", command=tv1.yview) # command means update the yaxis view of the widget
    treescrollx = Scrollbar(frame1, orient="horizontal", command=tv1.xview) # command means update the xaxis view of the widget
    tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set) # assign the scrollbars to the Treeview Widget
    treescrollx.pack(side="bottom", fill="x") # make the scrollbar fill the x axis of the Treeview widget
    treescrolly.pack(side="right", fill="y") # make the scrollbar fill the y axis of the Treeview widget


    def File_dialog():
        """This Function will open the file explorer and assign the chosen file path to label_file"""
        filename = filedialog.askopenfilename(initialdir="/",
                                            title="Select A File",
                                            filetype=(("xlsx files", "*.xlsx"),("All Files", "*.*")))
        label_file["text"] = filename
        return None


    def Load_excel_data():
        """If the file selected is valid this will load the file into the Treeview"""
        file_path = label_file["text"]
        try:
            excel_filename = r"{}".format(file_path)
            if excel_filename[-4:] == ".csv":
                df = pd.read_csv(excel_filename)
            else:
                df = pd.read_excel(excel_filename)

        except ValueError:
            messagebox.showerror("Information", "The file you have chosen is invalid")
            return None
        except FileNotFoundError:
            messagebox.showerror("Information", f"No such file as {file_path}")
            return None

        clear_data()
        tv1["column"] = list(df.columns)
        tv1["show"] = "headings"
        for column in tv1["columns"]:
            tv1.heading(column, text=column) # let the column heading = column name

        df_rows = df.to_numpy().tolist() # turns the dataframe into a list of lists
        for row in df_rows:
            tv1.insert("", "end", values=row) # inserts each list into the treeview. For parameters see https://docs.python.org/3/library/tkinter.ttk.html#tkinter.ttk.Treeview.insert
        return None


    def clear_data():
        tv1.delete(*tv1.get_children())
        return None



    outputw.mainloop()




#End of admin final window




#Admin credential checking 

def adminn():
    root.destroy()
    #Homepage creation 
    adcre = Tk()
    adcre.title('HOME PAGE')
    adcre['bg'] = '#29C5F6'
    adcre.resizable(False, False)  # This code helps to disable windows from resizing
    #setting window default size for home page
    window_height = 500
    window_width = 900

    screen_width = adcre.winfo_screenwidth()
    screen_height = adcre.winfo_screenheight()

    x_cordinate = int((screen_width/2) - (window_width/2))
    y_cordinate = int((screen_height/2) - (window_height/2))
    adcre.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))



    # Create an object for homepage

    instruction = Label(adcre, text = 'Please Enter Login Credentials', font=('Bahnschrift', 15, 'italic', 'bold'), bg ='#29C5F6',fg = 'Black')
    instruction.grid(columnspan=4, row=0)
    #Declaring Login credentials
    global password
    global username

    #Designing objects for login
    us = Label(adcre, text='USERNAME', font=('Arial', 15, 'bold'), bg ='#29C5F6').grid(row=1, column=0, pady=5)
    ps = Label(adcre, text='PASSWORD', font=('Arial', 15, 'bold'), bg ='#29C5F6').grid(row=2, column=0, pady=10)
    username = Entry(adcre, width=25, font=('Arial 20'))
    username.grid(row=1, column=1)
    password = Entry(adcre, show = '*', width=25, font=('Arial 20'))
    password.grid(row=2, column=1)

    def second():
        en1 = username.get()
        en2 = password.get()
            
        if en1 == 'username'  and en2 =='password':
            adcre.destroy()
            admin2()

           
            
        else:
            label1 = Label(adcre, text='Invalid Credentials')
            label1.grid(row=4, column=0)
            username.delete(0, END)
            password.delete(0, END)  




    Submit = Button(adcre, text = 'Submit',font=('Arial', 15), bg ='Red', width=10, border=8, command=second)
    Submit.grid(row=3,column=0, columnspan=2)


    adcre.mainloop()

    



#end of admin credential checking 


#Homepage creation 
root = Tk()
root.title('HOME PAGE')
root['bg'] = '#29C5F6'
root.resizable(False, False)  # This code helps to disable windows from resizing
#setting window default size for home page
window_height = 500
window_width = 900

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

x_cordinate = int((screen_width/2) - (window_width/2))
y_cordinate = int((screen_height/2) - (window_height/2))
root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))



# Create an object for homepage
title1 = Label(root, text = 'DEPARTMENT OF COMPUTER SCIENCE', font=('Pickwick', 20, 'bold' ), fg = '#000080', bg ='#29C5F6', justify=CENTER)
title1.pack()
title2 = Label(root, text = 'FEDERAL POLYTECHNIC BIDA', font=('Bahnschrift', 20), fg = 'Red', bg ='#29C5F6', justify=CENTER)
title2.pack()
title2 = Label(root, text = 'STUDENT ACADEMIC PEFORMANCE SYSTEM', font=('Bahnschrift', 20, 'bold'), fg = 'Black', bg ='#29C5F6',justify=CENTER)
title2.pack()
img = ImageTk.PhotoImage(Image.open("prediction4.png"))
label = Label(root, image = img, bg ='#29C5F6')
label.pack()

instruction = Label(root, text = 'Please Select Appropriate User', font=('Bahnschrift', 15, 'italic', 'bold'), bg ='#29C5F6',fg = 'Black')
instruction.pack()


#Form button creation 
def new():
    root.destroy() #closing hompage
    
    
    #Loading an existing workbook
 
    file = pathlib.Path('Prediction.xlsx')
    if file.exists ():
        file=load_workbook('Prediction.xlsx')
        sheet=file.active
    
    #creating a workbook for the form 
    else:
        file = openpyxl.Workbook()
        sheet = file.active
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 50
        sheet.column_dimensions['H'].width = 50
        
        
        # write given data to an excel spreadsheet
        # at particular location
        sheet.cell(row=1, column=1).value = "Name" 
        sheet.cell(row=1, column=2).value = "Course"
        sheet.cell(row=1, column=3).value = "Semester"
        sheet.cell(row=1, column=4).value = "Matric No."
        sheet.cell(row=1, column=5).value = "Phone Number"
        sheet.cell(row=1, column=6).value = "Email id"
        sheet.cell(row=1, column=7).value = "Address"
        sheet.cell(row=1, column=8).value = "Prediction Analysis"
        

    
    def excel():

            
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 40
            sheet.column_dimensions['G'].width = 50
            sheet.column_dimensions['H'].width = 50
            
            
            # write given data to an excel spreadsheet
            # at particular location
            sheet.cell(row=1, column=1).value = "Name" 
            sheet.cell(row=1, column=2).value = "Course"
            sheet.cell(row=1, column=3).value = "Semester"
            sheet.cell(row=1, column=4).value = "Matric No."
            sheet.cell(row=1, column=5).value = "Phone Number"
            sheet.cell(row=1, column=6).value = "Email id"
            sheet.cell(row=1, column=7).value = "Address"
            sheet.cell(row=1, column=8).value = "Prediction Analysis"
        
        


    # Function to set focus (cursor)
    def focus1(event):
        # set focus on the course_field box
        name_field.focus_set()

    def focus2(event):
        # set focus on the course_field box
        course_field.focus_set()
    
    
    # Function to set focus
    def focus3(event):
        # set focus on the sem_field box
        sem_field.focus_set()
    
    
    # Function to set focus
    def focus4(event):
        # set focus on the form_no_field box
        mat_no_field.focus_set()
    
    
    # Function to set focus
    def focus5(event):
        # set focus on the contact_no_field box
        phone_no_field.focus_set()
    
    
    # Function to set focus
    def focus6(event):
        # set focus on the email_id_field box
        email_id_field.focus_set()
    
    
    # Function to set focus
    def focus7(event):
        # set focus on the address_field box
        address_field.focus_set()
    
    # Function to set focus
    def focus8(event):
        # set focus on the address_field box
        Prediction.focus_set()
    # Function for clearing the
    # contents of text entry boxes
    def clear():
        
        # clear the content of text entry box
        name_field.delete(0, END)
        course_field.delete(0, END)
        sem_field.delete(0, END)
        mat_no_field.delete(0, END)
        phone_no_field.delete(0, END)
        email_id_field.delete(0, END)
        address_field.delete(0, END)
    
    
    # Function to take data from GUI
    # window and write to an excel file
    def insert():
        
        # if user not fill any entry
        # then print "empty input"
        if (name_field.get() == "" and
            course_field.get() == "" and
            sem_field.get() == "" and
            mat_no_field.get() == "" and
            phone_no_field.get() == "" and
            email_id_field.get() == "" and
            address_field.get() == ""):
                
            print("empty input")
    
        else:
    
            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = sheet.max_row
            current_column = sheet.max_column
    
            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location
            sheet.cell(row=current_row + 1, column=1).value = name_field.get()
            sheet.cell(row=current_row + 1, column=2).value = course_field.get()
            sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
            sheet.cell(row=current_row + 1, column=4).value = mat_no_field.get()
            sheet.cell(row=current_row + 1, column=5).value = phone_no_field.get()
            sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
            sheet.cell(row=current_row + 1, column=7).value = address_field.get()
            
            # save the file
    
    
            # set focus on the name_field box
            name_field.focus_set()
    
            # call the clear() function
            clear()
        root1.destroy()

        
        # Create questions as a list of dictionaries
        # The GUI will be generated automatically
        questions = [
            {"question": "1. Father’s Education Level?",
            "answers": ('',"No formal education", "Primary Education", "Secondary Education", "Graduate (OND/BSc. etc)", "Masters & Above")},
            {"question": "2. Mother’s Education Level?",
            "answers": ('',"No formal education", "Primary Education", "Secondary Education", "Graduate (OND/BSc. etc)", "Masters & Above")},
            {"question": "3. Marital Status?",
            "answers": ('',"Divorced", "Married", "Single")},
            {"question": "4. Parental income?",
            "answers": ('', "Low", "High", "Average")},
            {"question": "5. Gender?",
            "answers": ('', "Female", "Male")},
            {"question": "6. Parents’ involvement in academic activities e.g joint reading?",
            "answers": ('', "Low", "High", "Average")},
            {"question": "7. Age",
            "answers": ('',"Above 30", "Below 20", "20 - 30")},
            {"question": "8. Who do you live with?",
            "answers": ('', "Single Father", "Single Mother", "Relatives", "Alone", "Both parents")},
            {"question": "9. Who pays your bills?",
            "answers": ('', "Charity Organization", "Relatives", "Myself", "Brother / Sister", "Parent")},
            {"question": "10. What’s your plan after school?",
            "answers": ('', "I don't Know", "Get Married", "Get a Job", "Learn a Trade", "Build Myself")},
            {"question": "11. Reason for schooling?",
            "answers": ('', "No Reason", "Frustration", "Friends", "Parent", "Personal")},
            {"question": "12. Study habits?",
            "answers": ('', "I have no time to study", "I study only when I like", "I am lazy to study", "I study only it’s necessary ", "I am lazy to study")},
            {"question": "13. Report on assignments & classwork?",
            "answers": ('', "I get my  assignment done by myself", "Copy the assignments of friends", "I pay to get it done", "I do assignment only when its necessary", "I don't do assignment",)},
            {"question": "14.Availability of learning materials?",
            "answers": ('', "The student does not have any resources", "The student borrows resources from friends only when necessary", "The student has all necessary resources")},
            {"question": "15. What's your motivation to Study",
            "answers": ('', "To end up with a good job", "To learn", "To pass exams ")},
            {"question": "16.	How would you rate your level of accumulation?",
            "answers": ('', "Low", "High", "Low")},
            {"question": "17.	How would you rate your invovlement in class activities ",
            "answers": ('', "Not Active", "Very Active", "Active")},
            {"question": "18. What's your reason for absenteeism in school? ",
            "answers": ('', "Prefer not to say", "Tiredness", "Boredom", "Choice", "Illness")},
            {"question": "19. What's your level of absenteeism in school?",
            "answers": ('', "High", "Average", "Low")},
            {"question": "20. Previous Academic Performance?",
            "answers": ('', "Low", "Average", "High")},
            ]



        class App(Tk):
            
            def __init__(self):
                super().__init__()

                self.title("PREDICTION QUESTIONS")  # set the window title
                self.resizable(False, True)  # make window unresizable by width

                canv_frame = Frame(self)  # create the canvas frame
                # create the Canvas widget
                # highlightthickness=0 removes the black border when the canvas gets focus
                self.canv = Canvas(canv_frame, highlightthickness=0, height=500, width=800)
                # add scrolling when mouse wheel is rotated
                self.canv.bind_all("<MouseWheel>",
                                lambda event: self.canv.yview_scroll(-1 * (event.delta // 120), "units"))
                self.canv.pack(fill=BOTH, expand=YES, side=LEFT)  # pack the Canvas

                # Create a scrollbar
                # command=self.canv.yview tells the scrollbar to change the canvas yview
                # and canvas's yscrollcommand=self.yscrollbar.set tells the canvas to update
                # the scrollbar if canvas's yview is changed without it.
                self.yscrollbar = Scrollbar(canv_frame, command=self.canv.yview)
                self.canv["yscrollcommand"] = self.yscrollbar.set
                self.yscrollbar.pack(fill=Y, side=LEFT)  # pack the Scrollbar

                for question_id, question in enumerate(questions, 1):
                    qaframe = Frame(self.canv)  # create the question-answers (QA) frame
                    text = Text(qaframe, width=100, height=1.2, bg='orange', fg='white')  # create the Text widget for question
                    text.insert(END, question["question"])  # insert the question text there
                    text.pack(fill=X)  # pack the text widget
                    aframe = Frame(qaframe)  # create the answers frame
                    # Create the question variable and add it to the variables list
                    question_var = IntVar(self)
                    question["variable"] = question_var
                    # create the radiobuttons 
                    for answer_id, answer in enumerate(question["answers"]):
                        Radiobutton(aframe, variable=question_var, text=answer, value=answer_id).pack()
                    aframe.pack(fill=Y)  # pack the answers frame
                    self.canv.create_window(210, question_id * 175, window=qaframe)  # insert the QA frame into the Canvas
                canv_frame.pack(fill=BOTH, expand=YES)  # pack the canvas frame
                Button(self, text="Submit", command=self.submit, fg='Red', bg='Black', font=(12)).pack(fill=X)  # create the "Submit" button
                self.update()  # update everything to get the right scrollregion.
                self.canv.configure(scrollregion=self.canv.bbox("all"))  # set the canvas scrollregion
        
            def submit(self):
                sum_ = 0  # initially, the sum_ equals 0
                for question in questions:
                    sum_ += question["variable"].get()  # and then, we add all the questions answers
                if sum_>=80:
                    pout = 'High Performance'
                elif sum_>= 55:
                    pout = 'Average Performance'
                else:
                    pout = 'Low Performance'
                sheet.cell(row=current_row + 1, column=8).value =pout
                
                file.save("Prediction.xlsx")
                self.destroy()





        if __name__ == "__main__":  # if the App is not imported from another module,
            App().mainloop()  # create it and start the mainloop    

        
    
    # Driver code
    if __name__ == "__main__":
        
        # create a GUI window
        root1 = Tk()
    
        # set the background colour of GUI window
        root1.configure(background='light green')
    
        # set the title of GUI window
        root1.title("registration form")
    
        # set the configuration of GUI window
        window_height = 500
        window_width = 800

        screen_width = root1.winfo_screenwidth()
        screen_height = root1.winfo_screenheight()

        x_cordinate = int((screen_width/2) - (window_width/2))
        y_cordinate = int((screen_height/2) - (window_height/2))
        root1.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))

        excel()
    
        # create a Form label
        heading = Label(root1, text="ENTER THE FOLLOWING INFORMATION", bg="light green", font=('Bahnschrift', 20))
    
        # create a Name label
        name = Label(root1, text="Name", bg="light green", font=('Bahnschrift', 14))
    
        # create a Course label
        course = Label(root1, text="Course", bg="light green", font=('Bahnschrift', 14))
    
        # create a Semester label
        sem = Label(root1, text="Semester", bg="light green", font=('Bahnschrift', 14))
    
        # create a Form No. label
        mat_no = Label(root1, text="Matric No.", bg="light green", font=('Bahnschrift', 14))
    
        # create a Contact No. label
        phone_no = Label(root1, text="Phone No.", bg="light green", font=('Bahnschrift', 14))
    
        # create a Email id label
        email_id = Label(root1, text="Email id", bg="light green", font=('Bahnschrift', 14))
    
        # create a address label
        address = Label(root1, text="Address", bg="light green", font=('Bahnschrift', 14))
    
        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        heading.grid(row=0, column=1)
        name.grid(row=1, column=0)
        course.grid(row=2, column=0)
        sem.grid(row=3, column=0)
        mat_no.grid(row=4, column=0)
        phone_no.grid(row=5, column=0)
        email_id.grid(row=6, column=0)
        address.grid(row=7, column=0)
    
        # create a text entry box
        # for typing the information
        name_field = Entry(root1, font=(25))
        course_field = Entry(root1, font=(25))
        sem_field = Entry(root1, font=(25))
        mat_no_field = Entry(root1, font=(25))
        phone_no_field = Entry(root1, font=(25))
        email_id_field = Entry(root1, font=(25))
        address_field = Entry(root1, font=(25))
    
        # bind method of widget is used for
        # the binding the function with the events
    
        # whenever the enter key is pressed
        # then call the focus1 function
        name_field.bind("<Return>", focus1)
    
        # whenever the enter key is pressed
        # then call the focus2 function
        course_field.bind("<Return>", focus2)
    
        # whenever the enter key is pressed
        # then call the focus3 function
        sem_field.bind("<Return>", focus3)
    
        # whenever the enter key is pressed
        # then call the focus4 function
        mat_no_field.bind("<Return>", focus4)
    
        # whenever the enter key is pressed
        # then call the focus5 function
        phone_no_field.bind("<Return>", focus5)
    
        # whenever the enter key is pressed
        # then call the focus6 function
        email_id_field.bind("<Return>", focus6)


        # whenever the enter key is pressed
        # then call the focus6 function
        address_field.bind("<Return>", focus7)

    
        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        name_field.grid(row=1, column=1, ipadx="200", ipady="10", pady=5)
        course_field.grid(row=2, column=1, ipadx="200", ipady="10", pady=5)
        sem_field.grid(row=3, column=1, ipadx="200", ipady="10", pady=5)
        mat_no_field.grid(row=4, column=1, ipadx="200", ipady="10", pady=5)
        phone_no_field.grid(row=5, column=1, ipadx="200", ipady="10", pady=5)
        email_id_field.grid(row=6, column=1, ipadx="200", ipady="10", pady=5)
        address_field.grid(row=7, column=1, ipadx="200", ipady="10", pady=5 )
    
        # call excel function
        #Function to go back to home page
       
    

        # create a Submit Button and place into the root1 window

        submit = Button(root1, text="Submit", fg="White",font=('Arial', 15), 
                                 bg ='Green', width=20, border=8, command=insert)
        submit.grid(row=8, column=1)
        
    
        # start the GUI
        root1.mainloop()
    
    
admin = Button(root, text = 'Administrator',font=('Arial', 15), bg ='#5e1914', width=20, border=8, command=adminn)
admin.pack()

user = Button(root, text = 'User', font=('Arial', 15),  bg ='#29ab87', width=20, border=8, command = new,)
user.pack()



# Create a Label Widget to display the text or Image


root.mainloop()